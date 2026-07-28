"""Fonte automática: INSS — benefícios por município (EMPS/MPS).

XLSX anual nacional das Estatísticas Municipais da Previdência Social
(SÍNTESE/Dataprev): ben_municipios_especie_{ano}.xlsx (~3 MB, anos >= 2019).
Abas Qtd_dez{ano} (estoque de benefícios em dezembro) e Valor_Total_{ano}
(valor emitido no ano) — mesmas colunas A–M, header em 3 linhas mescladas,
dados a partir da linha cujo campo B é código IBGE de 7 dígitos.
REPLACE por (município, ano) com as 7 categorias-folha oficiais (subtotais
e Total ficam de fora — dupla contagem)."""
import io
from datetime import date

import requests

from app.services.ingestao_automatica.base import FonteAutomatica, ResumoIngestao, registrar
from app.services.ingestao_automatica.util import eh_nao_publicado

URL = "https://www.gov.br/previdencia/pt-br/assuntos/previdencia-social/arquivos/ben_municipios_especie_{ano}.xlsx"
INICIO_SERIE = 2019  # anos anteriores existem com nomes/paths antigos fora do padrão

# (índice 0-based na linha, nome da categoria) — só folhas mutuamente
# exclusivas: somam o Total (col 12); subtotais 3 e 10 ficam de fora.
CATEGORIAS: list[tuple[int, str]] = [
    (4, "Aposentadorias por idade"),
    (5, "Aposentadorias por invalidez"),
    (6, "Aposentadorias por tempo de contribuição"),
    (7, "Pensões por morte"),
    (8, "Auxílios"),
    (9, "Outros benefícios previdenciários"),
    (11, "Benefícios assistenciais"),
]


def parse_emps_aba(rows) -> dict[str, dict[str, float]]:
    """Linhas de uma aba do EMPS → {codigo_ibge: {categoria: valor}}.
    Linha de dados = campo B (índice 1) com código IBGE de 7 dígitos; o resto
    (headers mesclados, totais Brasil, rodapés) é ignorado."""
    out: dict[str, dict[str, float]] = {}
    for row in rows:
        codigo = str(row[1] if len(row) > 1 and row[1] is not None else "").strip()
        if not (codigo.isdigit() and len(codigo) == 7):
            continue
        vals: dict[str, float] = {}
        for idx, categoria in CATEGORIAS:
            v = row[idx] if len(row) > idx else None
            vals[categoria] = float(v) if v is not None and str(v).strip() != "" else 0.0
        out[codigo] = vals
    return out


def montar_registros(qtd_por_codigo, valor_por_codigo, alvo: dict[str, int], ano: int) -> list[dict]:
    """Casa Qtd × Valor por código IBGE dos municípios-alvo → dicts prontos
    para InssAnual(**d). Município ausente das DUAS abas fica de fora."""
    regs: list[dict] = []
    for codigo, mid in alvo.items():
        qtd = qtd_por_codigo.get(codigo)
        val = valor_por_codigo.get(codigo)
        if qtd is None and val is None:
            continue
        for _, categoria in CATEGORIAS:
            regs.append({
                "municipio_id": mid,
                "ano": ano,
                "categoria": categoria,
                "quantidade_beneficios": int((qtd or {}).get(categoria, 0.0)),
                "valor_anual": round((val or {}).get(categoria, 0.0), 2),
            })
    return regs


def achar_aba(sheetnames: list[str], prefixo: str) -> str | None:
    """Resolve aba por prefixo case-insensitive ('qtd', 'valor_total') — o
    sufixo varia entre anos (dez2024 vs dez24)."""
    for nome in sheetnames:
        if nome.lower().startswith(prefixo.lower()):
            return nome
    return None


def executar(db, municipios, anos=None, usuario_id=None, notificar=True, progresso=None) -> ResumoIngestao:
    import openpyxl

    from app.models.inss import InssAnual

    resumo = ResumoIngestao(dataset="inss")
    alvo = {str(m.codigo_ibge).strip(): m.id for m in municipios if m.codigo_ibge}
    for m in municipios:
        if not m.codigo_ibge:
            resumo.erros.append(f"{m.nome}/{m.estado}: sem codigo_ibge cadastrado")
            resumo.municipios_erro += 1
    if not alvo:
        return resumo

    ultimo_encerrado = date.today().year - 1
    anos_alvo = sorted({a for a in (anos or [ultimo_encerrado - 1, ultimo_encerrado]) if a >= INICIO_SERIE})
    mids_ok: set[int] = set()
    nao_publicados: list[str] = []

    for i, ano in enumerate(anos_alvo, start=1):
        if progresso:
            progresso(len(mids_ok), len(alvo), f"baixando EMPS {ano} ({i}/{len(anos_alvo)})")
        try:
            resp = requests.get(URL.format(ano=ano), timeout=(30, 300),
                                headers={"User-Agent": "Mozilla/5.0"})
            resp.raise_for_status()
        except requests.RequestException as exc:
            if eh_nao_publicado(exc):
                nao_publicados.append(str(ano))
            else:
                resumo.erros.append(f"EMPS {ano}: indisponível ({exc})")
            continue

        wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
        aba_qtd = achar_aba(wb.sheetnames, "qtd")
        aba_valor = achar_aba(wb.sheetnames, "valor_total")
        if not aba_qtd or not aba_valor:
            resumo.erros.append(f"EMPS {ano}: abas não reconhecidas ({wb.sheetnames}) — layout mudou?")
            continue
        qtd = parse_emps_aba(wb[aba_qtd].iter_rows(values_only=True))
        val = parse_emps_aba(wb[aba_valor].iter_rows(values_only=True))
        regs = montar_registros(qtd, val, alvo, ano)

        mids_do_ano = {r["municipio_id"] for r in regs}
        if mids_do_ano:
            db.query(InssAnual).filter(
                InssAnual.municipio_id.in_(mids_do_ano), InssAnual.ano == ano,
            ).delete(synchronize_session=False)
        for r in regs:
            db.add(InssAnual(**r))
        db.commit()
        resumo.linhas += len(regs)
        mids_ok |= mids_do_ano
        if progresso:
            progresso(len(mids_ok), len(alvo), f"EMPS {ano} gravado")

    if nao_publicados:
        anos_txt = ", ".join(nao_publicados)
        plural = "s" if len(nao_publicados) > 1 else ""
        resumo.erros.append(f"EMPS: ano{plural} {anos_txt} ainda não publicado{plural} pela Previdência")

    resumo.municipios_ok = len(mids_ok)
    faltantes = set(alvo.values()) - mids_ok
    resumo.municipios_erro += len(faltantes)
    if faltantes:
        nomes = {m.id: f"{m.nome}/{m.estado}" for m in municipios}
        for mid in sorted(faltantes):
            resumo.erros.append(f"{nomes.get(mid, mid)}: não encontrado no EMPS")
    return resumo


registrar(FonteAutomatica(
    key="inss",
    label="INSS (EMPS/Previdência)",
    fonte="EMPS — Estatísticas Municipais da Previdência Social (MPS/Dataprev): benefícios emitidos por município e categoria",
    executar=executar,
))
