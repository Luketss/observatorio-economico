"""Fonte com arquivo: IPS — Índice de Progresso Social (ipsbrasil.org.br).

O site é uma SPA sem URL estável de download (verificado 2026-07-27 e
2026-08-10) — o admin baixa o ips_brasil_municipios_{ano}.xlsx na UI do site
e envia pela tela de coletas; o blob trafega pelo banco (ingestao_arquivo)
porque API e worker não compartilham filesystem. Aceita também o CSV já
convertido (';', utf-8-sig). UPSERT por (municipio_id, ano): reenviar o
arquivo corrige dados. Reusa o COLUMN_MAP do CLI ingestao/carregar_ips."""
import csv
import io

from ingestao.carregar_ips import COLUMN_MAP
from app.services.ingestao_automatica.base import FonteAutomatica, registrar

COLUNAS_ESSENCIAIS = ("Código IBGE", "UF")


def _para_float(v) -> float | None:
    """Aceita número (XLSX) e string com vírgula decimal (CSV)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def ler_linhas(conteudo: bytes) -> list[dict]:
    """Bytes do upload → dicts header→valor. XLSX (zip 'PK') ou CSV ';'."""
    if conteudo[:2] == b"PK":
        return _ler_xlsx(conteudo)
    return _ler_csv(conteudo)


def _ler_xlsx(conteudo: bytes) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # o XLSX do site declara dimensão errada (A1) — sem isto só a 1ª célula sai
    ws.reset_dimensions()
    linhas = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(linhas, [])]
    return [dict(zip(header, row)) for row in linhas if not all(v is None for v in row)]


def _ler_csv(conteudo: bytes) -> list[dict]:
    return list(csv.DictReader(io.StringIO(conteudo.decode("utf-8-sig")), delimiter=";"))


def validar_headers(linhas: list[dict]) -> str | None:
    """None se ok; mensagem legível se o arquivo não parece o IPS nacional."""
    if not linhas:
        return "arquivo vazio ou sem linhas de dados — o IPS nacional tem ~5.570 municípios"
    headers = set(linhas[0].keys())
    faltando = [c for c in COLUNAS_ESSENCIAIS if c not in headers]
    if faltando or not headers & set(COLUMN_MAP):
        cols = ", ".join(f"'{c}'" for c in faltando) or "as colunas de métricas"
        return f"arquivo não parece ser o IPS nacional — colunas ausentes: {cols}"
    return None


def identificacao(row: dict) -> tuple[str | None, str, str]:
    """(codigo_ibge, nome sem o sufixo ' (UF)', UF em caixa alta)."""
    codigo = str(row.get("Código IBGE") or "").strip()
    nome = str(row.get("Município") or row.get("Municipio") or "").strip()
    if "(" in nome:
        nome = nome[: nome.index("(")].strip()
    uf = str(row.get("UF") or "").strip().upper()
    return (codigo or None, nome, uf)


def linha_para_kwargs(row: dict, municipio_id: int, ano: int) -> dict:
    """Kwargs prontos para IpsMunicipio(**kwargs) — mesma semântica do CLI."""
    kwargs = {"municipio_id": municipio_id, "ano": ano}
    kwargs["area_km2"] = _para_float(row.get("Área (km²)", row.get("Area (km2)")))
    populacao = _para_float(row.get("População 2022", row.get("Populacao 2022")))
    kwargs["populacao"] = int(populacao) if populacao is not None else None
    kwargs["pib_per_capita"] = _para_float(row.get("PIB per capita 2021"))
    for coluna, campo in COLUMN_MAP.items():
        if coluna in row and campo not in kwargs:
            kwargs[campo] = _para_float(row[coluna])
    return kwargs


def executar(db, municipios, anos=None, usuario_id=None, notificar=True,
             progresso=None, arquivo_id=None):
    """Carrega o arquivo IPS do blob para ips_municipio (upsert por
    município/ano). `municipios` e `notificar` são ignorados: o arquivo é
    nacional e a fonte não gera notificações. Falha dura = RuntimeError —
    o runner marca o job como 'erro' com a mensagem."""
    from sqlalchemy.dialects.postgresql import insert as pg_insert

    from app.models.ingestao_arquivo import IngestaoArquivo
    from app.models.ips import IpsMunicipio
    from app.services.ingestao_automatica.base import ResumoIngestao
    from ingestao.utils import obter_ou_criar_municipio

    ano = (anos or [None])[0]
    if not ano:
        raise RuntimeError("ano não informado — reenvie o arquivo pela tela de coletas")
    arq = db.get(IngestaoArquivo, arquivo_id) if arquivo_id else None
    if arq is None:
        raise RuntimeError("arquivo do upload não encontrado no banco — reenvie pela tela de coletas")

    if progresso:
        progresso(0, None, f"lendo {arq.nome}")
    try:
        linhas = ler_linhas(arq.conteudo)
    except Exception as exc:
        raise RuntimeError("arquivo inválido ou corrompido — baixe novamente do site e reenvie") from exc
    problema = validar_headers(linhas)
    if problema:
        raise RuntimeError(problema)

    resumo = ResumoIngestao(dataset="ips")
    total = len(linhas)
    colunas_upsert = None
    for i, row in enumerate(linhas, start=1):
        codigo, nome, uf = identificacao(row)
        if not nome or not uf:
            resumo.erros.append(f"linha {i}: sem município/UF — ignorada")
            continue
        municipio = obter_ou_criar_municipio(db, nome, uf, codigo)
        kwargs = linha_para_kwargs(row, municipio.id, ano)
        if colunas_upsert is None:
            colunas_upsert = [c for c in kwargs if c not in ("municipio_id", "ano")]
        stmt = pg_insert(IpsMunicipio).values(**kwargs).on_conflict_do_update(
            index_elements=["municipio_id", "ano"],
            set_={c: kwargs.get(c) for c in colunas_upsert},
        )
        db.execute(stmt)
        resumo.linhas += 1
        resumo.municipios_ok += 1
        if progresso and (i % 200 == 0 or i == total):
            progresso(i, total, f"IPS {ano}: {i}/{total} municípios")
    db.commit()

    db.delete(arq)  # blob cumpriu o papel — só sai depois do commit dos dados
    db.commit()
    return resumo


registrar(FonteAutomatica(
    key="ips",
    label="IPS — Índice de Progresso Social",
    fonte="IPS Brasil (ipsbrasil.org.br) — arquivo anual enviado pela tela de coletas",
    executar=executar,
    requer_arquivo=True,
))
